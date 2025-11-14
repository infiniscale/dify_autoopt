"""
结果采集模块 - Excel 导出器

Date: 2025-11-13
Author: backend-developer
Description: 将测试结果和统计数据导出为Excel格式报表
"""

from pathlib import Path
from typing import List, Optional
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from loguru import logger as loguru_logger

from ..utils.exceptions import ExportException
from .models import TestResult, PerformanceMetrics, ClassificationResult, PerformanceGrade


# 延迟初始化 logger，避免未初始化异常
def _get_safe_logger():
    """获取安全的 logger 实例"""
    try:
        from ..utils.logger import get_logger
        return get_logger(__name__)
    except Exception:
        # 如果 logger 未初始化，使用 loguru 默认 logger
        return loguru_logger.bind(name=__name__)


logger = _get_safe_logger()


class ExcelExporter:
    """
    Excel 数据导出器

    将测试结果导出为包含多个工作表的 Excel 文件:
    - Sheet1: 测试概览 (统计摘要)
    - Sheet2: 详细结果 (所有测试记录)
    - Sheet3: 性能分析 (按工作流分组)

    Example:
        >>> exporter = ExcelExporter()
        >>> exporter.export_results(results, "output/report.xlsx")
        >>> exporter.export_statistics(metrics, classification, "output/stats.xlsx")
    """

    # 样式常量
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    def __init__(self):
        """初始化导出器"""
        logger.info("ExcelExporter initialized")

    def export_results(
        self,
        results: List[TestResult],
        output_path: Path,
        include_stats: bool = True
    ) -> Path:
        """
        导出测试结果到 Excel

        Args:
            results: 测试结果列表
            output_path: 输出文件路径
            include_stats: 是否包含统计工作表

        Returns:
            输出文件的绝对路径

        Raises:
            ExportException: 导出失败时
        """
        try:
            if not results:
                raise ExportException("No results to export")

            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 移除默认工作表

            # 创建详细结果工作表
            self._create_details_sheet(wb, results)

            # 创建性能分析工作表
            self._create_performance_sheet(wb, results)

            # 如果需要统计,创建概览工作表
            if include_stats:
                # 计算基本统计
                from .data_collector import DataCollector
                from .classifier import ResultClassifier

                collector = DataCollector()
                for result in results:
                    collector.collect_result(result)

                metrics = collector.get_statistics()
                classifier = ResultClassifier()
                classification = classifier.classify_batch(results)

                self._create_overview_sheet(wb, metrics, classification)

            wb.save(output_path)
            logger.info(f"Exported {len(results)} results to {output_path}")

            return output_path.absolute()

        except ExportException:
            raise
        except Exception as e:
            logger.error(f"Export failed: {e}")
            raise ExportException(f"Failed to export results: {e}")

    def export_statistics(
        self,
        metrics: PerformanceMetrics,
        classification: ClassificationResult,
        output_path: Path
    ) -> Path:
        """
        导出统计数据到 Excel

        Args:
            metrics: 性能指标
            classification: 分类统计
            output_path: 输出文件路径

        Returns:
            输出文件的绝对路径

        Raises:
            ExportException: 导出失败时
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # 移除默认工作表

            # 创建概览工作表
            self._create_overview_sheet(wb, metrics, classification)

            wb.save(output_path)
            logger.info(f"Exported statistics to {output_path}")

            return output_path.absolute()

        except Exception as e:
            logger.error(f"Statistics export failed: {e}")
            raise ExportException(f"Failed to export statistics: {e}")

    def _create_overview_sheet(self, wb, metrics: PerformanceMetrics, classification: ClassificationResult):
        """创建概览工作表"""
        ws = wb.create_sheet("测试概览", 0)

        # 标题
        ws['A1'] = "测试概览报告"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # 执行统计
        ws['A5'] = "执行统计"
        ws['A5'].font = Font(bold=True, size=12)

        ws['A6'] = "总执行次数"
        ws['B6'] = metrics.total_executions

        ws['A7'] = "成功次数"
        ws['B7'] = metrics.successful_count

        ws['A8'] = "失败次数"
        ws['B8'] = metrics.failed_count

        ws['A9'] = "成功率"
        ws['B9'] = f"{metrics.success_rate:.2%}"

        # 性能统计
        ws['A11'] = "性能统计"
        ws['A11'].font = Font(bold=True, size=12)

        ws['A12'] = "平均执行时间"
        ws['B12'] = f"{metrics.avg_execution_time:.3f}s"

        ws['A13'] = "P50执行时间"
        ws['B13'] = f"{metrics.p50_execution_time:.3f}s"

        ws['A14'] = "P95执行时间"
        ws['B14'] = f"{metrics.p95_execution_time:.3f}s"

        ws['A15'] = "P99执行时间"
        ws['B15'] = f"{metrics.p99_execution_time:.3f}s"

        # 成本统计
        ws['A17'] = "成本统计"
        ws['A17'].font = Font(bold=True, size=12)

        ws['A18'] = "总Token数"
        ws['B18'] = f"{metrics.total_tokens:,}"

        ws['A19'] = "总成本"
        ws['B19'] = f"${metrics.total_cost:.2f}"

        ws['A20'] = "平均Token数"
        ws['B20'] = f"{metrics.avg_tokens_per_request:.1f}"

        # 性能分级
        ws['A22'] = "性能分级"
        ws['A22'].font = Font(bold=True, size=12)

        ws['A23'] = "优秀"
        ws['B23'] = f"{classification.excellent_count} ({classification.grade_distribution.get(PerformanceGrade.EXCELLENT, 0):.2f}%)"

        ws['A24'] = "良好"
        ws['B24'] = f"{classification.good_count} ({classification.grade_distribution.get(PerformanceGrade.GOOD, 0):.2f}%)"

        ws['A25'] = "一般"
        ws['B25'] = f"{classification.fair_count} ({classification.grade_distribution.get(PerformanceGrade.FAIR, 0):.2f}%)"

        ws['A26'] = "较差"
        ws['B26'] = f"{classification.poor_count} ({classification.grade_distribution.get(PerformanceGrade.POOR, 0):.2f}%)"

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

    def _create_details_sheet(self, wb, results: List[TestResult]):
        """创建详细结果工作表"""
        ws = wb.create_sheet("详细结果", 1)

        # 表头
        headers = ["工作流ID", "执行ID", "时间戳", "状态", "执行时间(s)", "Token数", "成本($)", "错误信息"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据行
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=result.workflow_id)
            ws.cell(row=row_num, column=2, value=result.execution_id)
            ws.cell(row=row_num, column=3, value=result.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=4, value=result.status.value)
            ws.cell(row=row_num, column=5, value=round(result.execution_time, 3))
            ws.cell(row=row_num, column=6, value=result.tokens_used)
            ws.cell(row=row_num, column=7, value=round(result.cost, 4))
            ws.cell(row=row_num, column=8, value=result.error_message or "")

        # 自动调整列宽
        self._auto_adjust_column_width(ws)

    def _create_performance_sheet(self, wb, results: List[TestResult]):
        """创建性能分析工作表"""
        ws = wb.create_sheet("性能分析", 2)

        # 表头
        headers = ["工作流ID", "执行次数", "成功率", "平均时间(s)", "P95时间(s)", "总Token", "总成本($)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # 按工作流分组
        workflow_stats = defaultdict(list)

        for result in results:
            workflow_stats[result.workflow_id].append(result)

        # 计算每个工作流的统计
        row_num = 2
        for workflow_id, wf_results in workflow_stats.items():
            total = len(wf_results)
            successful = sum(1 for r in wf_results if r.status.value == "success")
            success_rate = successful / total if total > 0 else 0

            exec_times = [r.execution_time for r in wf_results]
            avg_time = sum(exec_times) / len(exec_times) if exec_times else 0

            # P95
            sorted_times = sorted(exec_times)
            p95_index = int(0.95 * len(sorted_times))
            p95_time = sorted_times[p95_index] if sorted_times else 0

            total_tokens = sum(r.tokens_used for r in wf_results)
            total_cost = sum(r.cost for r in wf_results)

            ws.cell(row=row_num, column=1, value=workflow_id)
            ws.cell(row=row_num, column=2, value=total)
            ws.cell(row=row_num, column=3, value=f"{success_rate:.2%}")
            ws.cell(row=row_num, column=4, value=round(avg_time, 3))
            ws.cell(row=row_num, column=5, value=round(p95_time, 3))
            ws.cell(row=row_num, column=6, value=total_tokens)
            ws.cell(row=row_num, column=7, value=round(total_cost, 4))

            row_num += 1

        self._auto_adjust_column_width(ws)

    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
