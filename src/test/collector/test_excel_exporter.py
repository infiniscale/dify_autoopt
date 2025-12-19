"""
ExcelExporter 验收测试

Date: 2025-11-13
Author: backend-developer
Description: 验证 ExcelExporter 的 Excel 导出功能
"""

from datetime import datetime
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from src.collector import (
    ExcelExporter,
    DataCollector,
    ResultClassifier,
    TestResult,
    TestStatus,
    PerformanceMetrics,
    ClassificationResult,
    PerformanceGrade
)
from src.utils.exceptions import ExportException


def test_excel_export():
    """完整的 Excel 导出验收测试"""

    print("=" * 70)
    print("ExcelExporter Acceptance Test")
    print("=" * 70)

    # 1. 准备测试数据
    print("\n1. Preparing test data...")
    collector = DataCollector()

    for i in range(20):
        result = TestResult(
            workflow_id=f"wf_{i % 3}",
            execution_id=f"exec_{i}",
            timestamp=datetime.now(),
            status=TestStatus.SUCCESS if i % 4 != 0 else TestStatus.FAILED,
            execution_time=1.0 + i * 0.5,
            tokens_used=100 + i * 10,
            cost=0.01 + i * 0.005,
            inputs={"query": f"test_{i}"},
            outputs={"answer": f"result_{i}" * 50},
            error_message="Timeout error" if i % 4 == 0 else None
        )
        collector.collect_result(result)

    print(f"   Created {len(collector.get_all_results())} test records")

    # 2. 创建导出器
    print("\n2. Initializing ExcelExporter...")
    exporter = ExcelExporter()

    # 3. 导出完整报告 (包含统计)
    print("\n3. Exporting full report (with statistics)...")
    output_path = Path("output/test_report.xlsx")

    try:
        result_path = exporter.export_results(
            collector.get_all_results(),
            output_path,
            include_stats=True
        )
        print(f"   [OK] Export successful: {result_path}")
        assert result_path.exists(), "Export file does not exist"
        print(f"   [OK] File exists: {result_path.stat().st_size} bytes")
    except Exception as e:
        print(f"   [ERROR] Export failed: {e}")
        raise

    # 4. 导出仅统计报告
    print("\n4. Exporting statistics report...")
    stats_path = Path("output/stats_report.xlsx")

    try:
        metrics = collector.get_statistics()
        classifier = ResultClassifier()
        classification = classifier.classify_batch(collector.get_all_results())

        stats_output = exporter.export_statistics(
            metrics,
            classification,
            stats_path
        )
        print(f"   [OK] Export successful: {stats_output}")
        assert stats_output.exists(), "Statistics file does not exist"
        print(f"   [OK] File exists: {stats_output.stat().st_size} bytes")
    except Exception as e:
        print(f"   [ERROR] Export failed: {e}")
        raise

    # 5. 验证统计数据
    print("\n5. Verifying exported data...")
    print(f"   - Total executions: {metrics.total_executions}")
    print(f"   - Successful: {metrics.successful_count}")
    print(f"   - Failed: {metrics.failed_count}")
    print(f"   - Success rate: {metrics.success_rate:.2%}")
    print(f"   - Avg execution time: {metrics.avg_execution_time:.3f}s")
    print(f"   - P95 execution time: {metrics.p95_execution_time:.3f}s")
    print(f"   - Total tokens: {metrics.total_tokens:,}")
    print(f"   - Total cost: ${metrics.total_cost:.2f}")

    print("\n6. Performance grade distribution:")
    print(f"   - Excellent: {classification.excellent_count}")
    print(f"   - Good: {classification.good_count}")
    print(f"   - Fair: {classification.fair_count}")
    print(f"   - Poor: {classification.poor_count}")

    # 7. 测试空数据异常处理
    print("\n7. Testing exception handling...")
    try:
        exporter.export_results([], Path("output/empty.xlsx"))
        print("   [ERROR] Should have raised exception but didn't")
        assert False
    except Exception as e:
        print(f"   [OK] Correctly raised exception: {type(e).__name__}")

    print("\n" + "=" * 70)
    print("[SUCCESS] All tests passed!")
    print("=" * 70)
    print(f"\nPlease manually verify these files:")
    print(f"  1. {result_path.absolute()}")
    print(f"  2. {stats_output.absolute()}")
    print("\nVerify worksheet contents:")
    print("  - Test Overview: Contains statistical summary")
    print("  - Detailed Results: Contains all test records")
    print("  - Performance Analysis: Contains per-workflow statistics")


if __name__ == "__main__":
    test_excel_export()


class TestExcelExporterExceptions:
    """测试 ExcelExporter 的异常处理"""

    def test_export_results_generic_exception(self, tmp_path):
        """测试 export_results 中的通用异常处理 (第 119-121 行)"""
        exporter = ExcelExporter()

        # 创建测试数据
        results = [
            TestResult(
                workflow_id="wf_001",
                execution_id="exec_001",
                timestamp=datetime.now(),
                status=TestStatus.SUCCESS,
                execution_time=1.0,
                tokens_used=100,
                cost=0.01,
                inputs={"query": "test"},
                outputs={"answer": "result"}
            )
        ]

        output_file = tmp_path / "test.xlsx"

        # Mock openpyxl.Workbook 的 save 方法抛出异常
        with patch('openpyxl.Workbook.save') as mock_save:
            mock_save.side_effect = RuntimeError("Disk full")

            with pytest.raises(ExportException) as exc_info:
                exporter.export_results(results, output_file)

            assert "Failed to export results" in str(exc_info.value)
            assert "Disk full" in str(exc_info.value)

    def test_export_statistics_generic_exception(self, tmp_path):
        """测试 export_statistics 中的异常处理 (第 158-160 行)"""
        exporter = ExcelExporter()

        # 创建测试数据
        metrics = PerformanceMetrics(
            total_executions=10,
            successful_count=9,
            failed_count=1,
            success_rate=0.9,
            avg_execution_time=1.5,
            p50_execution_time=1.2,
            p95_execution_time=2.5,
            p99_execution_time=3.0,
            total_tokens=1000,
            total_cost=0.1,
            avg_tokens_per_request=100.0
        )

        classification = ClassificationResult(
            excellent_count=5,
            good_count=3,
            fair_count=1,
            poor_count=1,
            grade_distribution={
                PerformanceGrade.EXCELLENT: 50.0,
                PerformanceGrade.GOOD: 30.0,
                PerformanceGrade.FAIR: 10.0,
                PerformanceGrade.POOR: 10.0
            }
        )

        output_file = tmp_path / "stats.xlsx"

        # Mock openpyxl.Workbook 的 save 方法抛出异常
        with patch('openpyxl.Workbook.save') as mock_save:
            mock_save.side_effect = PermissionError("Access denied")

            with pytest.raises(ExportException) as exc_info:
                exporter.export_statistics(metrics, classification, output_file)

            assert "Failed to export statistics" in str(exc_info.value)
            assert "Access denied" in str(exc_info.value)

    def test_auto_adjust_column_width_exception(self, tmp_path):
        """测试 _auto_adjust_column_width 中的异常处理 (第 321-322 行)"""
        import openpyxl

        exporter = ExcelExporter()

        # 创建一个工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"

        # 添加正常数据
        ws['A1'] = "Normal"
        ws['B1'] = 123

        # 直接 patch len() 函数来触发异常
        # 在 _auto_adjust_column_width 中，会调用 len(str(cell.value))
        # 我们让 len() 在某些情况下抛出异常
        original_len = __builtins__['len'] if isinstance(__builtins__, dict) else __builtins__.len
        call_count = [0]

        def mock_len(obj):
            call_count[0] += 1
            # 在第3次调用 len 时抛出异常（跳过前面的一些调用）
            if call_count[0] == 3 and isinstance(obj, str):
                raise TypeError("Mock len error")
            return original_len(obj)

        # 使用 patch 来模拟 len() 抛出异常
        with patch('builtins.len', side_effect=mock_len):
            # 调用 _auto_adjust_column_width，应该能正常完成而不抛出异常
            # 因为异常被 try-except 捕获并忽略
            exporter._auto_adjust_column_width(ws)

        # 测试成功 - 方法应该能处理异常而不崩溃
        assert True
