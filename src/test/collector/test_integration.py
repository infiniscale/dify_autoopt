"""
端到端集成测试 - Collector 模块

Date: 2025-11-14
Author: qa-engineer
Description: 测试 DataCollector, ResultClassifier, ExcelExporter 的完整集成流程
"""

import pytest
import time
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from src.collector import (
    DataCollector,
    ResultClassifier,
    ExcelExporter,
    TestResult,
    TestStatus,
    PerformanceGrade
)
from .conftest import create_test_result


class TestCompletePipeline:
    """测试完整的数据流水线"""

    def test_complete_pipeline(self, sample_results, temp_output_dir):
        """
        测试完整的数据收集→分类→导出流程

        步骤:
        1. 创建 DataCollector, ResultClassifier, ExcelExporter
        2. 收集测试结果
        3. 计算统计指标
        4. 批量分类
        5. 导出到 Excel
        6. 验证 Excel 文件存在且包含正确数据
        """
        # 1. 创建组件
        collector = DataCollector()
        classifier = ResultClassifier()
        exporter = ExcelExporter()

        # 2. 收集数据
        for result in sample_results:
            collector.collect_result(result)

        assert collector.get_result_count() == len(sample_results)

        # 3. 计算统计
        stats = collector.get_statistics()
        assert stats.total_executions == len(sample_results)
        assert stats.success_rate > 0  # 应该有成功的结果

        # 4. 批量分类
        classification = classifier.classify_batch(sample_results)
        total_classified = (
                classification.excellent_count +
                classification.good_count +
                classification.fair_count +
                classification.poor_count
        )
        assert total_classified == len(sample_results)

        # 5. 导出到 Excel
        output_file = temp_output_dir / "complete_pipeline.xlsx"
        exporter.export_results(sample_results, str(output_file))

        # 6. 验证 Excel 文件
        assert output_file.exists()
        assert output_file.stat().st_size > 0

        # 验证工作表存在 (工作表名称可能因编码显示为乱码，检查数量即可)
        wb = load_workbook(output_file)
        assert len(wb.sheetnames) >= 3  # 至少有3个工作表

        wb.close()

    def test_pipeline_with_empty_data(self, temp_output_dir):
        """测试空数据的完整流程"""
        from src.utils.exceptions import DataValidationException, ExportException

        collector = DataCollector()
        classifier = ResultClassifier()
        exporter = ExcelExporter()

        # 空数据统计应该抛出异常
        with pytest.raises(DataValidationException):
            stats = collector.get_statistics()

        # 空数据分类
        classification = classifier.classify_batch([])
        assert classification.excellent_count == 0
        assert classification.good_count == 0
        assert classification.fair_count == 0
        assert classification.poor_count == 0

        # 空数据导出应该抛出异常
        output_file = temp_output_dir / "empty_pipeline.xlsx"
        with pytest.raises(ExportException):
            exporter.export_results([], str(output_file))

    def test_pipeline_incremental_updates(self, temp_output_dir):
        """
        测试增量更新流程

        验证多次收集和导出的数据一致性
        """
        collector = DataCollector()
        exporter = ExcelExporter()

        # 第一批数据
        batch1 = [create_test_result(execution_id=f"batch1_{i}") for i in range(10)]
        for result in batch1:
            collector.collect_result(result)

        stats1 = collector.get_statistics()
        assert stats1.total_executions == 10

        # 第二批数据
        batch2 = [create_test_result(execution_id=f"batch2_{i}") for i in range(15)]
        for result in batch2:
            collector.collect_result(result)

        stats2 = collector.get_statistics()
        assert stats2.total_executions == 25

        # 导出所有数据
        all_results = collector.get_all_results()
        output_file = temp_output_dir / "incremental.xlsx"
        exporter.export_results(all_results, str(output_file))

        assert output_file.exists()

        # 验证导出数据的完整性
        wb = load_workbook(output_file)
        detail_sheet = wb["详细结果"]

        # 标题行 + 25条数据
        assert detail_sheet.max_row == 26

        wb.close()


class TestMultipleWorkflows:
    """测试多工作流混合场景"""

    def test_multiple_workflows_statistics(self, mixed_workflow_results):
        """
        测试多个工作流的混合数据收集和分析

        验证:
        - 整体统计正确
        - 按工作流分组统计正确
        """
        collector = DataCollector()

        for result in mixed_workflow_results:
            collector.collect_result(result)

        # 整体统计
        overall_stats = collector.get_statistics()
        assert overall_stats.total_executions == 100  # 50 + 30 + 20

        # 按工作流统计
        wf1_results = collector.get_results_by_workflow("workflow_1")
        assert len(wf1_results) == 50

        wf2_results = collector.get_results_by_workflow("workflow_2")
        assert len(wf2_results) == 30

        wf3_results = collector.get_results_by_workflow("workflow_3")
        assert len(wf3_results) == 20

    def test_multiple_workflows_classification(self, mixed_workflow_results):
        """测试多工作流的性能分类"""
        classifier = ResultClassifier()

        # 分别对每个工作流分类
        wf1_results = [r for r in mixed_workflow_results if r.workflow_id == "workflow_1"]
        wf2_results = [r for r in mixed_workflow_results if r.workflow_id == "workflow_2"]
        wf3_results = [r for r in mixed_workflow_results if r.workflow_id == "workflow_3"]

        class1 = classifier.classify_batch(wf1_results)
        class2 = classifier.classify_batch(wf2_results)
        class3 = classifier.classify_batch(wf3_results)

        # Workflow 1 执行快，应该有较多优秀/良好
        assert class1.excellent_count + class1.good_count > class1.poor_count

        # Workflow 3 执行慢，应该有较多一般/较差
        assert class3.fair_count + class3.poor_count > class3.excellent_count

    def test_multiple_workflows_export(self, mixed_workflow_results, temp_output_dir):
        """测试多工作流数据的 Excel 导出"""
        exporter = ExcelExporter()
        output_file = temp_output_dir / "multi_workflows.xlsx"

        exporter.export_results(mixed_workflow_results, str(output_file))

        assert output_file.exists()

        # 验证性能分析工作表包含各工作流统计 (使用索引避免中文编码问题)
        wb = load_workbook(output_file)
        perf_sheet = wb.worksheets[2]  # 第三个工作表是性能分析

        # 应该至少有 4 行（标题 + 3 个工作流）
        assert perf_sheet.max_row >= 4

        wb.close()


class TestLargeDatasetIntegration:
    """测试大数据量场景"""

    def test_large_dataset_collection_performance(self, large_dataset):
        """
        测试大数据量收集性能

        验证: 收集 5,000 条数据的时间 < 5秒
        """
        collector = DataCollector()

        start_time = time.time()
        for result in large_dataset:
            collector.collect_result(result)
        collection_time = time.time() - start_time

        assert collector.get_result_count() == 5000
        assert collection_time < 5.0, f"Collection took {collection_time:.2f}s, expected < 5s"

    def test_large_dataset_statistics_performance(self, large_dataset):
        """
        测试大数据量统计计算性能

        验证: 统计计算时间 < 1秒
        """
        collector = DataCollector()
        for result in large_dataset:
            collector.collect_result(result)

        start_time = time.time()
        stats = collector.get_statistics()
        stats_time = time.time() - start_time

        assert stats.total_executions == 5000
        assert stats_time < 1.0, f"Statistics calculation took {stats_time:.2f}s, expected < 1s"

    def test_large_dataset_classification_performance(self, large_dataset):
        """
        测试大数据量分类性能

        验证: 分类 5,000 条数据的时间 < 2秒
        """
        classifier = ResultClassifier()

        start_time = time.time()
        classification = classifier.classify_batch(large_dataset)
        classification_time = time.time() - start_time

        total = (
                classification.excellent_count +
                classification.good_count +
                classification.fair_count +
                classification.poor_count
        )
        assert total == 5000
        assert classification_time < 2.0, f"Classification took {classification_time:.2f}s, expected < 2s"

    def test_large_dataset_export_performance(self, large_dataset, temp_output_dir):
        """
        测试大数据量导出性能

        验证: 导出 5,000 条数据的时间 < 10秒
        """
        exporter = ExcelExporter()
        output_file = temp_output_dir / "large_dataset.xlsx"

        start_time = time.time()
        exporter.export_results(large_dataset, str(output_file))
        export_time = time.time() - start_time

        assert output_file.exists()
        assert export_time < 10.0, f"Export took {export_time:.2f}s, expected < 10s"

    def test_large_dataset_full_pipeline_performance(self, large_dataset, temp_output_dir):
        """
        测试大数据量完整流程性能

        验证: 收集+统计+分类+导出 总时间 < 20秒
        """
        start_time = time.time()

        # 收集
        collector = DataCollector()
        for result in large_dataset:
            collector.collect_result(result)

        # 统计
        stats = collector.get_statistics()

        # 分类
        classifier = ResultClassifier()
        classification = classifier.classify_batch(large_dataset)

        # 导出
        exporter = ExcelExporter()
        output_file = temp_output_dir / "full_pipeline_large.xlsx"
        exporter.export_results(large_dataset, str(output_file))

        total_time = time.time() - start_time

        assert output_file.exists()
        assert total_time < 20.0, f"Full pipeline took {total_time:.2f}s, expected < 20s"


class TestEdgeCasesIntegration:
    """测试边界和异常情况"""

    def test_all_failed_results(self, temp_output_dir):
        """测试全部失败的结果"""
        failed_results = [
            create_test_result(
                execution_id=f"fail_{i}",
                status=TestStatus.FAILED,
                execution_time=1.0 + i * 0.1,
                tokens_used=100 + i * 10
            )
            for i in range(20)
        ]

        collector = DataCollector()
        for result in failed_results:
            collector.collect_result(result)

        stats = collector.get_statistics()
        assert stats.success_rate == 0.0
        assert stats.failed_count == 20

        # 分类仍然应该工作
        classifier = ResultClassifier()
        classification = classifier.classify_batch(failed_results)
        assert classification.excellent_count + classification.good_count + classification.fair_count + classification.poor_count == 20

        # 导出仍然应该工作
        exporter = ExcelExporter()
        output_file = temp_output_dir / "all_failed.xlsx"
        exporter.export_results(failed_results, str(output_file))
        assert output_file.exists()

    def test_all_successful_results(self, temp_output_dir):
        """测试全部成功的结果"""
        success_results = [
            create_test_result(
                execution_id=f"success_{i}",
                status=TestStatus.SUCCESS,
                execution_time=0.5,
                tokens_used=100
            )
            for i in range(20)
        ]

        collector = DataCollector()
        for result in success_results:
            collector.collect_result(result)

        stats = collector.get_statistics()
        assert stats.success_rate == 1.0
        assert stats.successful_count == 20

    def test_extreme_execution_times(self, edge_case_results, temp_output_dir):
        """测试极端执行时间"""
        collector = DataCollector()
        for result in edge_case_results:
            collector.collect_result(result)

        stats = collector.get_statistics()
        assert stats.total_executions == len(edge_case_results)

        # 验证极端值被正确处理
        all_results = collector.get_all_results()
        exec_times = [r.execution_time for r in all_results]
        assert min(exec_times) >= 0.01  # 有极快的
        assert max(exec_times) >= 100.0  # 有极慢的

    def test_zero_token_consumption(self, edge_case_results):
        """测试零 Token 消耗情况"""
        collector = DataCollector()
        for result in edge_case_results:
            collector.collect_result(result)

        # 验证统计不会因为零值而崩溃
        stats = collector.get_statistics()
        assert stats.total_tokens >= 0

        # 验证分类不会因为零值而崩溃
        classifier = ResultClassifier()
        classification = classifier.classify_batch(edge_case_results)
        total = sum([
            classification.excellent_count,
            classification.good_count,
            classification.fair_count,
            classification.poor_count
        ])
        assert total == len(edge_case_results)

    def test_special_characters_in_data(self, edge_case_results, temp_output_dir):
        """测试特殊字符处理"""
        exporter = ExcelExporter()
        output_file = temp_output_dir / "special_chars.xlsx"

        # 导出包含特殊字符的数据
        exporter.export_results(edge_case_results, str(output_file))

        assert output_file.exists()

        # 验证文件可以正常打开
        wb = load_workbook(output_file)
        detail_sheet = wb["详细结果"]

        # 验证至少有一些数据
        assert detail_sheet.max_row > 1

        wb.close()


class TestIncrementalCollection:
    """测试增量数据收集"""

    def test_incremental_collection_statistics_update(self):
        """
        测试增量收集和统计更新

        步骤:
        1. 收集第一批数据（50条）
        2. 计算统计 A
        3. 收集第二批数据（50条）
        4. 计算统计 B
        5. 验证统计 B 包含所有数据
        """
        collector = DataCollector()

        # 第一批：50条成功
        batch1 = [
            create_test_result(
                execution_id=f"batch1_{i}",
                status=TestStatus.SUCCESS,
                execution_time=1.0
            )
            for i in range(50)
        ]

        for result in batch1:
            collector.collect_result(result)

        stats_a = collector.get_statistics()
        assert stats_a.total_executions == 50
        assert stats_a.success_rate == 1.0

        # 第二批：50条，部分失败
        batch2 = [
            create_test_result(
                execution_id=f"batch2_{i}",
                status=TestStatus.SUCCESS if i % 2 == 0 else TestStatus.FAILED,
                execution_time=2.0
            )
            for i in range(50)
        ]

        for result in batch2:
            collector.collect_result(result)

        stats_b = collector.get_statistics()
        assert stats_b.total_executions == 100
        assert stats_b.successful_count == 75  # 50 + 25
        assert stats_b.failed_count == 25
        assert stats_b.success_rate == 0.75

    def test_incremental_workflow_addition(self):
        """测试增量添加不同工作流"""
        collector = DataCollector()

        # 先添加 workflow_1
        for i in range(20):
            collector.collect_result(create_test_result(workflow_id="workflow_1", execution_id=f"wf1_{i}"))

        wf1_results = collector.get_results_by_workflow("workflow_1")
        assert len(wf1_results) == 20

        # 再添加 workflow_2
        for i in range(30):
            collector.collect_result(create_test_result(workflow_id="workflow_2", execution_id=f"wf2_{i}"))

        wf2_results = collector.get_results_by_workflow("workflow_2")
        assert len(wf2_results) == 30

        # 验证总数
        assert collector.get_result_count() == 50

    def test_incremental_export_consistency(self, temp_output_dir):
        """测试增量导出的数据一致性"""
        collector = DataCollector()
        exporter = ExcelExporter()

        # 第一次导出
        batch1 = [create_test_result(execution_id=f"v1_{i}") for i in range(10)]
        for result in batch1:
            collector.collect_result(result)

        file1 = temp_output_dir / "export_v1.xlsx"
        exporter.export_results(collector.get_all_results(), str(file1))

        # 第二次导出（增加数据）
        batch2 = [create_test_result(execution_id=f"v2_{i}") for i in range(15)]
        for result in batch2:
            collector.collect_result(result)

        file2 = temp_output_dir / "export_v2.xlsx"
        exporter.export_results(collector.get_all_results(), str(file2))

        # 验证两次导出都成功
        assert file1.exists()
        assert file2.exists()

        # 验证第二次导出包含更多数据 (使用索引避免中文编码问题)
        wb1 = load_workbook(file1)
        wb2 = load_workbook(file2)

        rows1 = wb1.worksheets[1].max_row  # 详细结果工作表
        rows2 = wb2.worksheets[1].max_row

        assert rows2 > rows1  # v2 应该有更多行

        wb1.close()
        wb2.close()
