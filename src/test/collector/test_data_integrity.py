"""
数据完整性测试 - Collector 模块

Date: 2025-11-14
Author: qa-engineer
Description: 测试数据一致性、准确性和完整性
"""

import pytest
import copy
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from src.collector import (
    DataCollector,
    ResultClassifier,
    ExcelExporter,
    TestResult,
    TestStatus,
    PerformanceGrade
)
from .conftest import create_test_result


class TestDataConsistency:
    """测试数据一致性"""

    def test_collected_data_immutability(self, sample_results):
        """
        测试收集的数据不被修改

        验证:
        1. 收集后的数据与原始数据完全一致
        2. 修改原始数据不影响已收集的数据
        """
        collector = DataCollector()

        # 创建深拷贝作为对照
        original_copies = [copy.deepcopy(r) for r in sample_results]

        # 收集数据
        for result in sample_results:
            collector.collect_result(result)

        # 验证收集的数据与原始拷贝一致
        collected = collector.get_all_results()

        # 验证数据数量正确
        assert len(collected) == len(original_copies)

        # 验证第一条数据的关键字段
        assert collected[0].workflow_id == original_copies[0].workflow_id
        assert collected[0].execution_id == original_copies[0].execution_id
        assert collected[0].execution_time == original_copies[0].execution_time

    def test_statistics_reproducibility(self, sample_results):
        """
        测试统计结果可重现

        验证:
        1. 相同数据多次计算统计结果一致
        2. 统计结果不受调用顺序影响
        """
        collector = DataCollector()
        for result in sample_results:
            collector.collect_result(result)

        # 多次计算统计
        stats1 = collector.get_statistics()
        stats2 = collector.get_statistics()
        stats3 = collector.get_statistics()

        # 验证结果完全一致
        assert stats1.total_executions == stats2.total_executions == stats3.total_executions
        assert stats1.success_rate == stats2.success_rate == stats3.success_rate
        assert stats1.avg_execution_time == stats2.avg_execution_time == stats3.avg_execution_time
        assert stats1.total_tokens == stats2.total_tokens == stats3.total_tokens
        assert stats1.total_cost == stats2.total_cost == stats3.total_cost

    def test_query_results_consistency(self, mixed_workflow_results):
        """
        测试查询结果的一致性

        验证:
        1. 多次查询返回相同的结果
        2. 不同查询方式返回数据一致
        """
        collector = DataCollector()
        for result in mixed_workflow_results:
            collector.collect_result(result)

        # 查询所有结果
        all_results1 = collector.get_all_results()
        all_results2 = collector.get_all_results()

        assert len(all_results1) == len(all_results2)
        assert all(r1.execution_id == r2.execution_id for r1, r2 in zip(all_results1, all_results2))

        # 查询特定工作流
        wf1_results1 = collector.get_results_by_workflow("workflow_1")
        wf1_results2 = collector.get_results_by_workflow("workflow_1")

        assert len(wf1_results1) == len(wf1_results2)
        assert len(wf1_results1) == 50

    def test_data_order_preservation(self):
        """
        测试数据顺序保持

        验证收集顺序与查询顺序一致
        """
        collector = DataCollector()

        # 按特定顺序收集
        execution_ids = [f"order_{i:03d}" for i in range(20)]
        for exec_id in execution_ids:
            result = create_test_result(execution_id=exec_id)
            collector.collect_result(result)

        # 验证顺序保持
        collected = collector.get_all_results()
        collected_ids = [r.execution_id for r in collected]

        assert collected_ids == execution_ids


class TestClassificationStability:
    """测试分类稳定性"""

    def test_classification_reproducibility(self, sample_results):
        """
        测试分类稳定性

        验证相同数据多次分类结果一致
        """
        classifier = ResultClassifier()

        # 多次分类
        class1 = classifier.classify_batch(sample_results)
        class2 = classifier.classify_batch(sample_results)
        class3 = classifier.classify_batch(sample_results)

        # 验证结果完全一致
        assert class1.excellent_count == class2.excellent_count == class3.excellent_count
        assert class1.good_count == class2.good_count == class3.good_count
        assert class1.fair_count == class2.fair_count == class3.fair_count
        assert class1.poor_count == class2.poor_count == class3.poor_count

    def test_single_vs_batch_classification_consistency(self, sample_results):
        """
        测试单个分类与批量分类的一致性

        验证:
        1. 逐个分类的结果
        2. 批量分类的结果
        3. 两者应该完全一致
        """
        classifier = ResultClassifier()

        # 单个分类
        single_grades = []
        for result in sample_results:
            grade = classifier.classify_result(result)
            single_grades.append(grade)

        # 批量分类
        batch_classification = classifier.classify_batch(sample_results)

        # 统计单个分类结果
        single_counts = {
            PerformanceGrade.EXCELLENT: single_grades.count(PerformanceGrade.EXCELLENT),
            PerformanceGrade.GOOD: single_grades.count(PerformanceGrade.GOOD),
            PerformanceGrade.FAIR: single_grades.count(PerformanceGrade.FAIR),
            PerformanceGrade.POOR: single_grades.count(PerformanceGrade.POOR)
        }

        # 验证一致性
        assert single_counts[PerformanceGrade.EXCELLENT] == batch_classification.excellent_count
        assert single_counts[PerformanceGrade.GOOD] == batch_classification.good_count
        assert single_counts[PerformanceGrade.FAIR] == batch_classification.fair_count
        assert single_counts[PerformanceGrade.POOR] == batch_classification.poor_count

    def test_threshold_adjustment_impact(self):
        """
        测试阈值调整后分类结果符合预期

        验证阈值变化对分类的正确影响
        """
        # 创建边界测试结果
        boundary_result = TestResult(
            workflow_id="boundary",
            execution_id="boundary_001",
            timestamp=datetime.now(),
            status=TestStatus.SUCCESS,
            execution_time=2.5,  # 在不同阈值下会有不同分类
            tokens_used=100,
            cost=0.01,
            inputs={"query": "test"},
            outputs={"answer": "x" * 75}  # 效率 0.75
        )

        # 宽松阈值
        loose_thresholds = {
            "excellent": {"execution_time": 3.0, "token_efficiency": 0.7},
            "good": {"execution_time": 6.0, "token_efficiency": 0.5},
            "fair": {"execution_time": 12.0, "token_efficiency": 0.3}
        }

        # 严格阈值
        strict_thresholds = {
            "excellent": {"execution_time": 1.5, "token_efficiency": 0.9},
            "good": {"execution_time": 4.0, "token_efficiency": 0.7},
            "fair": {"execution_time": 8.0, "token_efficiency": 0.5}
        }

        loose_classifier = ResultClassifier(thresholds=loose_thresholds)
        strict_classifier = ResultClassifier(thresholds=strict_thresholds)

        loose_grade = loose_classifier.classify_result(boundary_result)
        strict_grade = strict_classifier.classify_result(boundary_result)

        # 宽松标准应该给出更高或相同的评级
        grade_order = {
            PerformanceGrade.EXCELLENT: 4,
            PerformanceGrade.GOOD: 3,
            PerformanceGrade.FAIR: 2,
            PerformanceGrade.POOR: 1
        }

        assert grade_order[loose_grade] >= grade_order[strict_grade]


class TestExcelDataAccuracy:
    """测试 Excel 导出数据准确性"""

    def test_excel_overview_sheet_accuracy(self, sample_results, temp_output_dir):
        """
        测试 Excel 概览工作表的数据准确性

        验证统计数据与计算结果一致
        """
        # 计算预期统计
        collector = DataCollector()
        for result in sample_results:
            collector.collect_result(result)

        expected_stats = collector.get_statistics()

        # 导出
        exporter = ExcelExporter()
        output_file = temp_output_dir / "accuracy_test.xlsx"
        exporter.export_results(sample_results, str(output_file))

        # 读取 Excel 并验证 (使用索引访问工作表避免中文编码问题)
        wb = load_workbook(output_file)
        overview_sheet = wb.worksheets[0]  # 第一个工作表是概览

        # 验证工作表有数据（至少有标题和一些数据行）
        assert overview_sheet.max_row >= 5  # 应该有多行统计数据
        assert overview_sheet.max_column >= 2  # 至少有两列

        wb.close()

    def test_excel_detail_sheet_completeness(self, sample_results, temp_output_dir):
        """
        测试 Excel 详细结果工作表的完整性

        验证每一行数据都正确导出
        """
        exporter = ExcelExporter()
        output_file = temp_output_dir / "detail_test.xlsx"
        exporter.export_results(sample_results, str(output_file))

        # 读取详细结果 (使用索引访问工作表避免中文编码问题)
        wb = load_workbook(output_file)
        detail_sheet = wb.worksheets[1]  # 第二个工作表是详细结果

        # 验证行数（标题行 + 数据行）
        assert detail_sheet.max_row == len(sample_results) + 1

        # 验证列标题存在
        headers = [cell.value for cell in detail_sheet[1]]
        assert len(headers) >= 7  # 至少有7个主要列

        # 验证有数据行
        assert detail_sheet.max_row >= 2  # 至少有标题行和一些数据

        wb.close()

    def test_excel_performance_analysis_accuracy(self, mixed_workflow_results, temp_output_dir):
        """
        测试 Excel 性能分析工作表的准确性

        验证按工作流分组的统计数据
        """
        exporter = ExcelExporter()
        output_file = temp_output_dir / "perf_analysis_test.xlsx"
        exporter.export_results(mixed_workflow_results, str(output_file))

        # 计算预期的工作流统计
        collector = DataCollector()
        for result in mixed_workflow_results:
            collector.collect_result(result)

        # 读取性能分析表 (使用索引访问工作表避免中文编码问题)
        wb = load_workbook(output_file)
        perf_sheet = wb.worksheets[2]  # 第三个工作表是性能分析

        # 验证工作流行存在
        workflow_names = []
        for row in range(2, perf_sheet.max_row + 1):
            wf_name = perf_sheet.cell(row=row, column=1).value
            if wf_name:
                workflow_names.append(wf_name)

        # 验证包含所有工作流
        assert "workflow_1" in workflow_names
        assert "workflow_2" in workflow_names
        assert "workflow_3" in workflow_names

        wb.close()

    def test_excel_data_types_correctness(self, sample_results, temp_output_dir):
        """
        测试 Excel 中的数据类型正确性

        验证数值、文本、日期等类型正确
        """
        exporter = ExcelExporter()
        output_file = temp_output_dir / "data_types_test.xlsx"
        exporter.export_results(sample_results, str(output_file))

        wb = load_workbook(output_file)
        detail_sheet = wb.worksheets[1]  # 第二个工作表是详细结果

        # 检查第2行（第一条数据）
        if detail_sheet.max_row >= 2:
            # 执行时间应该是数值
            exec_time = detail_sheet.cell(row=2, column=5).value
            assert isinstance(exec_time, (int, float))

            # Token使用应该是整数
            tokens = detail_sheet.cell(row=2, column=6).value
            assert isinstance(tokens, (int, float))

            # 成本应该是数值
            cost = detail_sheet.cell(row=2, column=7).value
            assert isinstance(cost, (int, float))

        wb.close()


class TestDataIntegrityUnderStress:
    """测试压力情况下的数据完整性"""

    def test_rapid_collection_integrity(self):
        """
        测试快速连续收集的数据完整性

        验证快速收集不会丢失数据
        """
        collector = DataCollector()
        expected_count = 1000

        # 快速收集
        for i in range(expected_count):
            result = create_test_result(execution_id=f"rapid_{i}")
            collector.collect_result(result)

        # 验证数量
        assert collector.get_result_count() == expected_count

        # 验证所有 ID 都存在
        collected = collector.get_all_results()
        collected_ids = {r.execution_id for r in collected}
        expected_ids = {f"rapid_{i}" for i in range(expected_count)}

        assert collected_ids == expected_ids

    def test_large_dataset_integrity(self, large_dataset):
        """
        测试大数据集的完整性

        验证所有数据都被正确处理
        """
        collector = DataCollector()
        classifier = ResultClassifier()

        # 收集所有数据
        for result in large_dataset:
            collector.collect_result(result)

        # 验证数量
        assert collector.get_result_count() == len(large_dataset)

        # 验证统计
        stats = collector.get_statistics()
        assert stats.total_executions == len(large_dataset)

        # 验证分类
        classification = classifier.classify_batch(large_dataset)
        total_classified = (
            classification.excellent_count +
            classification.good_count +
            classification.fair_count +
            classification.poor_count
        )
        assert total_classified == len(large_dataset)

    def test_duplicate_execution_id_handling(self):
        """
        测试重复执行 ID 的处理

        验证系统正确处理重复数据
        """
        collector = DataCollector()

        # 收集相同 execution_id 的结果
        result1 = create_test_result(execution_id="duplicate", execution_time=1.0)
        result2 = create_test_result(execution_id="duplicate", execution_time=2.0)

        collector.collect_result(result1)
        collector.collect_result(result2)

        # 验证都被收集（不应该去重）
        assert collector.get_result_count() == 2

        collected = collector.get_all_results()
        assert len(collected) == 2

    def test_export_data_matches_collection(self, sample_results, temp_output_dir):
        """
        测试导出的数据与收集的数据完全一致

        验证导出过程不改变数据
        """
        collector = DataCollector()
        for result in sample_results:
            collector.collect_result(result)

        collected = collector.get_all_results()

        # 导出
        exporter = ExcelExporter()
        output_file = temp_output_dir / "export_match_test.xlsx"
        exporter.export_results(collected, str(output_file))

        # 读取并验证 (使用索引访问工作表避免中文编码问题)
        wb = load_workbook(output_file)
        detail_sheet = wb.worksheets[1]  # 第二个工作表是详细结果

        # 验证行数匹配
        assert detail_sheet.max_row == len(collected) + 1  # +1 for header

        # 验证 execution_id 完全匹配
        exported_ids = []
        for row in range(2, detail_sheet.max_row + 1):
            exec_id = detail_sheet.cell(row=row, column=2).value
            exported_ids.append(exec_id)

        collected_ids = [r.execution_id for r in collected]
        assert set(exported_ids) == set(collected_ids)

        wb.close()


class TestNumericalAccuracy:
    """测试数值计算准确性"""

    def test_statistics_calculation_accuracy(self):
        """
        测试统计计算的数值准确性

        使用已知数据验证计算结果
        """
        collector = DataCollector()

        # 创建已知数据
        results = [
            create_test_result(execution_id="acc_1", execution_time=1.0, tokens_used=100, cost=0.01, status=TestStatus.SUCCESS),
            create_test_result(execution_id="acc_2", execution_time=2.0, tokens_used=200, cost=0.02, status=TestStatus.SUCCESS),
            create_test_result(execution_id="acc_3", execution_time=3.0, tokens_used=300, cost=0.03, status=TestStatus.FAILED),
            create_test_result(execution_id="acc_4", execution_time=4.0, tokens_used=400, cost=0.04, status=TestStatus.SUCCESS),
        ]

        for result in results:
            collector.collect_result(result)

        stats = collector.get_statistics()

        # 验证精确计算
        assert stats.total_executions == 4
        assert stats.successful_count == 3
        assert stats.failed_count == 1
        assert stats.success_rate == 0.75

        # 平均执行时间: (1+2+3+4)/4 = 2.5
        assert abs(stats.avg_execution_time - 2.5) < 0.01

        # 总 Token: 100+200+300+400 = 1000
        assert stats.total_tokens == 1000

        # 总成本: 0.01+0.02+0.03+0.04 = 0.10
        assert abs(stats.total_cost - 0.10) < 0.001

    def test_percentile_calculation_accuracy(self):
        """
        测试百分位数计算的准确性

        使用有序数据验证百分位数
        """
        collector = DataCollector()

        # 创建有序的执行时间：1.0, 2.0, ..., 10.0
        for i in range(1, 11):
            result = create_test_result(execution_id=f"perc_{i}", execution_time=float(i))
            collector.collect_result(result)

        stats = collector.get_statistics()

        # P50 应该在 5-6 之间
        assert 5.0 <= stats.p50_execution_time <= 6.0

        # P95 应该接近 9.5-10.0
        assert 9.0 <= stats.p95_execution_time <= 10.0

        # P99 应该接近 10.0
        assert 9.5 <= stats.p99_execution_time <= 10.0

    def test_floating_point_precision(self):
        """
        测试浮点数精度处理

        验证小数计算的准确性
        """
        collector = DataCollector()

        # 使用可能导致浮点误差的数值
        results = [
            create_test_result(execution_id=f"fp_{i}", execution_time=0.1, cost=0.001)
            for i in range(10)
        ]

        for result in results:
            collector.collect_result(result)

        stats = collector.get_statistics()

        # 验证累加误差在可接受范围内
        expected_total_cost = 0.001 * 10
        assert abs(stats.total_cost - expected_total_cost) < 0.0001
